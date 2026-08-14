import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_workbook():
    # 1. Load our processed analytical data
    df = pd.read_csv("data/processed/fact_economic_indicators.csv")
    
    # 2. Setup the openpyxl workbook
    wb = openpyxl.Workbook()
    
    # --- SHEET 1: Data Summary & Pivot Simulation ---
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Clean styling definitions
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11, color="333333")
    font_summary_total = Font(name="Segoe UI", size=11, bold=True, color="000000")
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_total = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    
    # Risk Fills (Conditional formatting style)
    fill_stable = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # light green
    fill_watch = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # light yellow
    fill_elevated = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # light red
    
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    border_double_bottom = Border(
        bottom=Side(style='double', color='000000'),
        top=Side(style='thin', color='DDDDDD')
    )

    # Title Block
    ws_summary["A1"] = "Global Economic & Risk Intelligence - Executive Summary (2024)"
    ws_summary["A1"].font = font_title
    ws_summary.row_dimensions[1].height = 30
    
    # Write Table Headers
    headers = [
        "Country Code", "Country Name", "GDP Growth (%)", "Inflation (%)", 
        "Unemployment (%)", "Govt Debt (% GDP)", "Risk Score", "Risk Category"
    ]
    
    row_idx = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[row_idx].height = 24
    
    # Filter dataset for year 2024 to create a management summary
    df_2024 = df[df["Year"] == 2024].copy()
    
    # Write data lines
    row_idx = 4
    for idx, row in df_2024.iterrows():
        ws_summary.cell(row=row_idx, column=1, value=row["CountryCode"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=2, value=row["CountryName"]).alignment = Alignment(horizontal="left")
        
        # Numeric values
        ws_summary.cell(row=row_idx, column=3, value=round(row["GDP_Growth"], 2)).number_format = '0.00"%"'
        ws_summary.cell(row=row_idx, column=4, value=round(row["Inflation"], 2)).number_format = '0.00"%"'
        ws_summary.cell(row=row_idx, column=5, value=round(row["Unemployment"], 2)).number_format = '0.00"%"'
        ws_summary.cell(row=row_idx, column=6, value=round(row["Govt_Debt"], 2)).number_format = '0.00"%"'
        
        # Risk Score
        risk_score_cell = ws_summary.cell(row=row_idx, column=7, value=round(row["CountryRiskScore"], 2))
        risk_score_cell.number_format = '0.0'
        risk_score_cell.alignment = Alignment(horizontal="right")
        
        # Risk Category with explicit background colors
        cat_cell = ws_summary.cell(row=row_idx, column=8, value=row["RiskCategory"])
        cat_cell.alignment = Alignment(horizontal="center")
        if row["RiskCategory"] == "Stable":
            cat_cell.fill = fill_stable
        elif row["RiskCategory"] == "Watch":
            cat_cell.fill = fill_watch
        else:
            cat_cell.fill = fill_elevated
            
        # Apply fonts and thin borders
        for col_idx in range(1, 9):
            c = ws_summary.cell(row=row_idx, column=col_idx)
            if col_idx != 8: # don't overwrite risk color fill
                c.font = font_data
            c.border = border_thin
            
        ws_summary.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Write average/summary total row
    ws_summary.cell(row=row_idx, column=1, value="AVG").font = font_summary_total
    ws_summary.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
    
    # Formulas for averages
    for c_idx, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G')]:
        formula = f"=AVERAGE({letter}4:{letter}{row_idx-1})"
        cell = ws_summary.cell(row=row_idx, column=c_idx, value=formula)
        cell.font = font_summary_total
        cell.border = border_double_bottom
        cell.fill = fill_total
        if c_idx == 7:
            cell.number_format = '0.0'
        else:
            cell.number_format = '0.00"%"'
            
    # Apply total row styles for remaining cells
    ws_summary.cell(row=row_idx, column=2, value="Global Benchmarks").font = font_summary_total
    ws_summary.cell(row=row_idx, column=2).fill = fill_total
    ws_summary.cell(row=row_idx, column=2).border = border_double_bottom
    ws_summary.cell(row=row_idx, column=8, value="").border = border_double_bottom
    ws_summary.cell(row=row_idx, column=8).fill = fill_total
    ws_summary.row_dimensions[row_idx].height = 22
    
    # Auto-adjust column widths for premium feel
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # --- SHEET 2: Full Data Validation ---
    ws_data = wb.create_sheet(title="All Indicators Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    # Header block
    ws_data.cell(row=1, column=1, value="CountryCode").font = font_header
    ws_data.cell(row=1, column=1).fill = fill_header
    
    # Quick export of the raw Fact table structure
    # Header
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(row=1, column=c_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            
    wb.save("excel/tracker.xlsx")
    print("Excel tracker sheet successfully created at excel/tracker.xlsx")

if __name__ == "__main__":
    create_excel_workbook()
